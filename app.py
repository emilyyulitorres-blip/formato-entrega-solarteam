import base64
from io import BytesIO
from pathlib import Path
from datetime import datetime, date

import pandas as pd
import qrcode
import streamlit as st
import streamlit.components.v1 as components

from openpyxl import Workbook, load_workbook

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image
)


# ======================================================
# CONFIGURACIÓN
# ======================================================

st.set_page_config(
    page_title="Entrega y Recepción de Equipos",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

PRIMARY = "#102A56"
GREEN = "#D6E53F"

EXCEL_FILE = Path("registro_entregas.xlsx")
PDF_DIR = Path("pdf_generados")

PDF_DIR.mkdir(exist_ok=True)


UNIDADES = [
    "Unidad",
    "Metro",
    "Rollo",
    "Caja",
    "Kit",
    "Juego",
    "Par",
    "Global",
    "Otro"
]


# ======================================================
# ESTILOS
# ======================================================

st.markdown(
    """
    <style>

    [data-testid="stSidebar"] {
        background: linear-gradient(
            180deg,
            #061B38 0%,
            #102A56 100%
        );
    }

    [data-testid="stSidebar"] * {
        color: white;
    }

    [data-testid="stSidebar"] img {
        background: transparent;
        padding: 6px;
    }

    .block-container {
        padding-top: 1rem;
        padding-left: 1.5rem;
        padding-right: 1.5rem;
        max-width: 100%;
    }

    .main-title {
        font-size: 34px;
        line-height: 1.2;
        font-weight: 900;
        color: #102A56;
        margin: 0;
        padding: 0;
    }

    .subtitle {
        font-size: 15px;
        color: #667085;
        margin-top: 8px;
        margin-bottom: 18px;
    }

    .card {
        background: white;
        border: 1px solid #E5E7EB;
        border-radius: 14px;
        padding: 18px;
        box-shadow: 0 1px 4px rgba(16,42,86,0.06);
    }

    .section-title {
        color: #169B62;
        font-weight: 800;
        font-size: 17px;
        margin-bottom: 12px;
    }

    .sidebar-btn {
        display: block;
        padding: 13px 14px;
        margin: 10px 0;
        border-radius: 9px;
        font-weight: 700;
        background: rgba(255,255,255,0.06);
    }

    .sidebar-btn-active {
        background: #169B62;
    }

    .help-box {
        border: 1px solid rgba(255,255,255,0.35);
        border-radius: 10px;
        padding: 12px;
        margin-top: 80px;
        font-size: 13px;
    }

    .history-card {
        background: white;
        border: 1px solid #E5E7EB;
        border-radius: 14px;
        padding: 16px;
        margin-top: 15px;
    }

    .stButton > button {
        border-radius: 9px;
        font-weight: 700;
        height: 42px;
    }

    div[data-testid="stDownloadButton"] button {
        background-color: #169B62;
        color: white;
        border: none;
        border-radius: 9px;
        font-weight: 800;
        height: 42px;
    }

    div[data-testid="stDownloadButton"] button:hover {
        background-color: #11814F;
        color: white;
    }

    </style>
    """,
    unsafe_allow_html=True
)


# ======================================================
# FUNCIONES EXCEL
# ======================================================

def obtener_siguiente_documento():

    if not EXCEL_FILE.exists():
        return "ER-ST-001"

    try:

        wb = load_workbook(EXCEL_FILE)

        if "Registro General" not in wb.sheetnames:
            return "ER-ST-001"

        ws = wb["Registro General"]

        numeros = []

        for row in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            documento = row[0]

            if (
                documento
                and str(documento).startswith("ER-ST-")
            ):

                try:

                    numero = int(
                        str(documento).replace(
                            "ER-ST-",
                            ""
                        )
                    )

                    numeros.append(numero)

                except Exception:
                    pass

        siguiente = (
            max(numeros) + 1
            if numeros
            else 1
        )

        return f"ER-ST-{siguiente:03d}"

    except Exception:

        return "ER-ST-001"


def crear_excel_si_no_existe():

    if EXCEL_FILE.exists():
        return

    wb = Workbook()

    ws_general = wb.active

    ws_general.title = "Registro General"

    ws_general.append([
        "N° Documento",
        "Fecha",
        "Hora",
        "Destino",
        "Proyecto",
        "Transportista",
        "Entregado por",
        "Recibido por",
        "Equipo responsable",
        "Observaciones",
        "Archivo PDF"
    ])


    ws_materiales = wb.create_sheet(
        "Detalle Materiales"
    )

    ws_materiales.append([
        "N° Documento",
        "Material / Equipo",
        "Cantidad",
        "Unidad",
        "Entregado",
        "Recibido"
    ])


    wb.save(EXCEL_FILE)


def guardar_excel(
    data,
    materiales,
    pdf_name
):

    crear_excel_si_no_existe()

    wb = load_workbook(EXCEL_FILE)


    ws_general = wb["Registro General"]

    ws_general.append([

        data["numero_documento"],

        data["fecha"],

        data["hora"],

        data["destino"],

        data["proyecto"],

        data["transportista"],

        data["entregado_por"],

        data["recibido_por"],

        data["equipo_responsable"],

        data["observaciones"],

        pdf_name

    ])


    ws_materiales = wb[
        "Detalle Materiales"
    ]


    for material in materiales:

        ws_materiales.append([

            data["numero_documento"],

            material.get(
                "Material / Equipo",
                ""
            ),

            material.get(
                "Cantidad",
                ""
            ),

            material.get(
                "Unidad",
                ""
            ),

            (
                "Sí"
                if material.get(
                    "Entregado",
                    False
                )
                else "No"
            ),

            (
                "Sí"
                if material.get(
                    "Recibido",
                    False
                )
                else "No"
            )

        ])


    wb.save(EXCEL_FILE)


def cargar_historial():

    if not EXCEL_FILE.exists():

        return pd.DataFrame()


    try:

        return pd.read_excel(

            EXCEL_FILE,

            sheet_name="Registro General"

        )


    except Exception:

        return pd.DataFrame()


# ======================================================
# FUNCIONES GENERALES
# ======================================================

def logo_b64():

    logo_path = Path("logo.png")


    if not logo_path.exists():

        return ""


    return base64.b64encode(

        logo_path.read_bytes()

    ).decode("utf-8")


def normalizar_materiales(materiales):

    materiales_validos = []


    for material in materiales:

        nombre = str(

            material.get(

                "Material / Equipo",

                ""

            )

        ).strip()


        if (
            nombre
            and nombre.lower() != "none"
        ):

            materiales_validos.append(
                material
            )


    return materiales_validos


def generar_qr(data):

    texto = (

        f"N° Documento: "
        f"{data['numero_documento']}\n"

        f"Proyecto: "
        f"{data['proyecto']}\n"

        f"Destino: "
        f"{data['destino']}\n"

        f"Transportista: "
        f"{data['transportista']}\n"

        f"Fecha: "
        f"{data['fecha']}"

    )


    qr = qrcode.QRCode(

        box_size=3,

        border=2

    )


    qr.add_data(texto)

    qr.make(fit=True)


    imagen = qr.make_image(

        fill_color="black",

        back_color="white"

    )


    buffer = BytesIO()


    imagen.save(

        buffer,

        format="PNG"

    )


    buffer.seek(0)


    return buffer


# ======================================================
# GENERAR PDF
# ======================================================

def generar_pdf(
    data,
    materiales
):

    buffer = BytesIO()


    doc = SimpleDocTemplate(

        buffer,

        pagesize=A4,

        rightMargin=1.3 * cm,

        leftMargin=1.3 * cm,

        topMargin=1.1 * cm,

        bottomMargin=1.1 * cm

    )


    styles = getSampleStyleSheet()


    styles.add(

        ParagraphStyle(

            name="TituloPDF",

            parent=styles["Title"],

            fontName="Helvetica-Bold",

            fontSize=14,

            alignment=1,

            leading=18

        )

    )


    styles.add(

        ParagraphStyle(

            name="TextoPDF",

            parent=styles["Normal"],

            fontName="Helvetica",

            fontSize=8,

            leading=10

        )

    )


    styles.add(

        ParagraphStyle(

            name="SeccionPDF",

            parent=styles["Normal"],

            fontName="Helvetica-Bold",

            fontSize=9,

            textColor=colors.white

        )

    )


    story = []


    logo_path = Path("logo.png")


    if logo_path.exists():

        logo = Image(

            str(logo_path),

            width=4.1 * cm,

            height=2.0 * cm

        )

    else:

        logo = Paragraph(

            "SOLARTEAM",

            styles["TituloPDF"]

        )


    encabezado = Table(

        [[

            logo,

            Paragraph(

                "FORMULARIO DE ENTREGA Y"

                "<br/>"

                "RECEPCIÓN DE EQUIPOS",

                styles["TituloPDF"]

            ),

            Paragraph(

                f"<b>N° DOCUMENTO</b>"

                f"<br/>"

                f"{data['numero_documento']}",

                styles["TextoPDF"]

            )

        ]],

        colWidths=[

            5.0 * cm,

            8.4 * cm,

            4.2 * cm

        ]

    )


    encabezado.setStyle(

        TableStyle([

            (

                "BOX",

                (0, 0),

                (-1, -1),

                0.8,

                colors.HexColor("#9AA4B2")

            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "MIDDLE"

            ),

            (

                "ALIGN",

                (0, 0),

                (-1, -1),

                "CENTER"

            ),

            (

                "PADDING",

                (0, 0),

                (-1, -1),

                7

            )

        ])

    )


    story.append(encabezado)

    story.append(

        Spacer(

            1,

            8

        )

    )


    qr_img = Image(

        generar_qr(data),

        width=2.4 * cm,

        height=2.4 * cm

    )


    info_fecha = Table(

        [[

            Paragraph(

                f"<b>Fecha:</b> "

                f"{data['fecha']}",

                styles["TextoPDF"]

            ),

            Paragraph(

                f"<b>Hora:</b> "

                f"{data['hora']}",

                styles["TextoPDF"]

            ),

            qr_img

        ]],

        colWidths=[

            7.0 * cm,

            7.0 * cm,

            3.4 * cm

        ]

    )


    story.append(info_fecha)

    story.append(

        Spacer(

            1,

            8

        )

    )


    titulo_datos = Table(

        [[

            Paragraph(

                "DATOS GENERALES",

                styles["SeccionPDF"]

            )

        ]],

        colWidths=[

            17.6 * cm

        ]

    )


    titulo_datos.setStyle(

        TableStyle([

            (

                "BACKGROUND",

                (0, 0),

                (-1, -1),

                colors.HexColor(PRIMARY)

            ),

            (

                "PADDING",

                (0, 0),

                (-1, -1),

                5

            )

        ])

    )


    story.append(titulo_datos)


    datos = [

        [

            "Destino / Área:",

            data["destino"],

            "Proyecto asignado:",

            data["proyecto"]

        ],

        [

            "Transportista:",

            data["transportista"],

            "",

            ""

        ],

        [

            "Entregado por:",

            data["entregado_por"],

            "",

            ""

        ],

        [

            "Recibido por:",

            data["recibido_por"],

            "Equipo responsable:",

            data["equipo_responsable"]

        ],

        [

            "Observaciones:",

            (
                data["observaciones"]

                if data["observaciones"]

                else "Sin observaciones."
            ),

            "",

            ""

        ]

    ]


    tabla_datos = Table(

        datos,

        colWidths=[

            3.2 * cm,

            6.0 * cm,

            3.0 * cm,

            5.4 * cm

        ]

    )


    tabla_datos.setStyle(

        TableStyle([

            (

                "GRID",

                (0, 0),

                (-1, -1),

                0.45,

                colors.HexColor("#BFC7D5")

            ),

            (

                "BACKGROUND",

                (0, 0),

                (0, -1),

                colors.HexColor("#F1F5F9")

            ),

            (

                "BACKGROUND",

                (2, 0),

                (2, -1),

                colors.HexColor("#F1F5F9")

            ),

            (

                "FONTNAME",

                (0, 0),

                (0, -1),

                "Helvetica-Bold"

            ),

            (

                "FONTNAME",

                (2, 0),

                (2, -1),

                "Helvetica-Bold"

            ),

            (

                "FONTSIZE",

                (0, 0),

                (-1, -1),

                8

            ),

            (

                "SPAN",

                (1, 1),

                (3, 1)

            ),

            (

                "SPAN",

                (1, 2),

                (3, 2)

            ),

            (

                "SPAN",

                (1, 4),

                (3, 4)

            ),

            (

                "PADDING",

                (0, 0),

                (-1, -1),

                5

            )

        ])

    )


    story.append(tabla_datos)

    story.append(

        Spacer(

            1,

            10

        )

    )


    titulo_materiales = Table(

        [[

            Paragraph(

                "DETALLE DE MATERIALES / EQUIPOS",

                styles["SeccionPDF"]

            )

        ]],

        colWidths=[

            17.6 * cm

        ]

    )


    titulo_materiales.setStyle(

        TableStyle([

            (

                "BACKGROUND",

                (0, 0),

                (-1, -1),

                colors.HexColor(PRIMARY)

            ),

            (

                "PADDING",

                (0, 0),

                (-1, -1),

                5

            )

        ])

    )


    story.append(titulo_materiales)


    detalle = [[

        "No.",

        "Material / Equipo",

        "Cantidad",

        "Unidad",

        "Entregado",

        "Recibido"

    ]]


    for i, material in enumerate(

        materiales,

        start=1

    ):

        detalle.append([

            str(i),

            Paragraph(

                str(

                    material.get(

                        "Material / Equipo",

                        ""

                    )

                ),

                styles["TextoPDF"]

            ),

            str(

                material.get(

                    "Cantidad",

                    ""

                )

            ),

            str(

                material.get(

                    "Unidad",

                    ""

                )

            ),

            (
                "X"

                if material.get(

                    "Entregado",

                    False

                )

                else ""
            ),

            (
                "X"

                if material.get(

                    "Recibido",

                    False

                )

                else ""
            )

        ])


    tabla_materiales = Table(

        detalle,

        colWidths=[

            1.0 * cm,

            8.1 * cm,

            2.2 * cm,

            2.3 * cm,

            2.0 * cm,

            2.0 * cm

        ],

        repeatRows=1

    )


    tabla_materiales.setStyle(

        TableStyle([

            (

                "GRID",

                (0, 0),

                (-1, -1),

                0.45,

                colors.HexColor("#BFC7D5")

            ),

            (

                "BACKGROUND",

                (0, 0),

                (-1, 0),

                colors.HexColor("#F1F5F9")

            ),

            (

                "FONTNAME",

                (0, 0),

                (-1, 0),

                "Helvetica-Bold"

            ),

            (

                "FONTSIZE",

                (0, 0),

                (-1, -1),

                8

            ),

            (

                "ALIGN",

                (0, 0),

                (0, -1),

                "CENTER"

            ),

            (

                "ALIGN",

                (2, 1),

                (-1, -1),

                "CENTER"

            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "MIDDLE"

            ),

            (

                "PADDING",

                (0, 0),

                (-1, -1),

                5

            )

        ])

    )


    story.append(tabla_materiales)

    story.append(

        Spacer(

            1,

            18

        )

    )


    firmas = Table(

        [

            [

                "ENTREGA",

                "RECEPCIÓN"

            ],

            [

                "",

                ""

            ],

            [

                data["entregado_por"],

                data["recibido_por"]

            ],

            [

                "",

                data["equipo_responsable"]

            ],

            [

                "Nombre y firma",

                "Nombre y firma"

            ]

        ],

        colWidths=[

            8.8 * cm,

            8.8 * cm

        ],

        rowHeights=[

            0.6 * cm,

            2.0 * cm,

            0.5 * cm,

            0.5 * cm,

            0.5 * cm

        ]

    )


    firmas.setStyle(

        TableStyle([

            (

                "GRID",

                (0, 0),

                (-1, -1),

                0.45,

                colors.HexColor("#BFC7D5")

            ),

            (

                "BACKGROUND",

                (0, 0),

                (-1, 0),

                colors.HexColor(PRIMARY)

            ),

            (

                "TEXTCOLOR",

                (0, 0),

                (-1, 0),

                colors.white

            ),

            (

                "FONTNAME",

                (0, 0),

                (-1, 0),

                "Helvetica-Bold"

            ),

            (

                "ALIGN",

                (0, 0),

                (-1, -1),

                "CENTER"

            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "MIDDLE"

            ),

            (

                "FONTSIZE",

                (0, 0),

                (-1, -1),

                8

            )

        ])

    )


    story.append(firmas)


    doc.build(story)


    buffer.seek(0)


    return buffer


# ======================================================
# VISTA PREVIA HTML
# ======================================================

def pdf_preview_html(
    data,
    materiales
):

    logo_string = logo_b64()


    if logo_string:

        logo_html = (

            f'<img '

            f'src="data:image/png;base64,{logo_string}" '

            f'style="width:145px; height:auto;">'

        )

    else:

        logo_html = (

            '<div style="'

            'font-size:24px;'

            'font-weight:800;'

            'color:#D6E53F;'

            '">'

            'SOLARTEAM'

            '</div>'

        )


    filas = ""


    for i, material in enumerate(

        materiales,

        start=1

    ):

        entregado = (

            "✓"

            if material.get(

                "Entregado",

                False

            )

            else ""

        )


        recibido = (

            "✓"

            if material.get(

                "Recibido",

                False

            )

            else ""

        )


        filas += f"""

        <tr>

            <td>{i}</td>

            <td>
                {material.get("Material / Equipo", "")}
            </td>

            <td>
                {material.get("Cantidad", "")}
            </td>

            <td>
                {material.get("Unidad", "")}
            </td>

            <td class="check">
                {entregado}
            </td>

            <td class="check">
                {recibido}
            </td>

        </tr>

        """


    if not filas:

        filas = """

        <tr>

            <td colspan="6">

                Sin materiales registrados

            </td>

        </tr>

        """


    html = f"""

    <!DOCTYPE html>

    <html>

    <head>

    <meta charset="UTF-8">


    <style>


    * {{

        box-sizing: border-box;

    }}


    body {{

        margin: 0;

        padding: 0;

        font-family:

            Arial,

            Helvetica,

            sans-serif;

        background: transparent;

    }}


    .preview-box {{

        background: white;

        border: 1px solid #E5E7EB;

        border-radius: 14px;

        padding: 20px;

    }}


    .pdf-page {{

        background: white;

        border: 1px solid #CBD5E1;

        padding: 20px;

        min-height: 720px;

    }}


    .pdf-header {{

        display: grid;

        grid-template-columns:

            150px

            1fr

            180px;

        gap: 15px;

        align-items: center;

        border-bottom:

            1px solid #94A3B8;

        padding-bottom: 15px;

    }}


    .logo-box {{

        display: flex;

        align-items: center;

        justify-content: center;

    }}


    .pdf-title {{

        font-size: 20px;

        line-height: 1.6;

        font-weight: 800;

        color: #071B38;

        text-align: center;

    }}


    .doc-number {{

        border: 1px solid #94A3B8;

        padding: 16px;

        text-align: center;

        font-weight: 800;

        font-size: 15px;

        line-height: 1.8;

    }}


    .fecha-row {{

        display: flex;

        justify-content: space-between;

        align-items: center;

        padding: 14px 0;

        font-size: 13px;

    }}


    .qr-box {{

        width: 75px;

        height: 75px;

        border: 1px solid #D0D5DD;

        display: flex;

        align-items: center;

        justify-content: center;

        font-size: 12px;

    }}


    .pdf-section {{

        background: #102A56;

        color: white;

        font-weight: 800;

        padding: 7px 10px;

        font-size: 13px;

        margin-top: 10px;

    }}


    table {{

        width: 100%;

        border-collapse: collapse;

        font-size: 12px;

    }}


    th,

    td {{

        border: 1px solid #D0D5DD;

        padding: 7px;

    }}


    th {{

        background: #F2F4F7;

        font-weight: 700;

    }}


    .check {{

        color: #169B62;

        font-weight: 900;

        text-align: center;

    }}


    .firmas {{

        margin-top: 18px;

    }}


    .firma-space {{

        height: 70px;

        text-align: center;

        vertical-align: bottom;

    }}


    .footer {{

        border-top:

            2px solid #D6E53F;

        text-align: center;

        font-size: 12px;

        padding-top: 10px;

        margin-top: 18px;

    }}


    </style>


    </head>


    <body>


    <div class="preview-box">


    <div class="pdf-page">


    <div class="pdf-header">


    <div class="logo-box">

        {logo_html}

    </div>


    <div class="pdf-title">

        FORMULARIO DE ENTREGA

        <br>

        Y

        <br>

        RECEPCIÓN DE EQUIPOS

    </div>


    <div class="doc-number">

        N° DOCUMENTO

        <br>

        {data["numero_documento"]}

    </div>


    </div>


    <div class="fecha-row">


    <div>

        <b>Fecha:</b>

        {data["fecha"]}

    </div>


    <div>

        <b>Hora:</b>

        {data["hora"]}

    </div>


    <div class="qr-box">

        QR

    </div>


    </div>


    <div class="pdf-section">

        DATOS GENERALES

    </div>


    <table>


    <tr>

        <td>

            <b>Destino / Área:</b>

        </td>

        <td>

            {data["destino"]}

        </td>

        <td>

            <b>Proyecto:</b>

        </td>

        <td>

            {data["proyecto"]}

        </td>

    </tr>


    <tr>

        <td>

            <b>Transportista:</b>

        </td>

        <td colspan="3">

            {data["transportista"]}

        </td>

    </tr>


    <tr>

        <td>

            <b>Entregado por:</b>

        </td>

        <td colspan="3">

            {data["entregado_por"]}

        </td>

    </tr>


    <tr>

        <td>

            <b>Recibido por:</b>

        </td>

        <td>

            {data["recibido_por"]}

        </td>

        <td>

            <b>Equipo responsable:</b>

        </td>

        <td>

            {data["equipo_responsable"]}

        </td>

    </tr>


    <tr>

        <td>

            <b>Observaciones:</b>

        </td>

        <td colspan="3">

            {data["observaciones"]}

        </td>

    </tr>


    </table>


    <div class="pdf-section">

        DETALLE DE MATERIALES / EQUIPOS

    </div>


    <table>


    <tr>

        <th>No.</th>

        <th>Material / Equipo</th>

        <th>Cantidad</th>

        <th>Unidad</th>

        <th>Entregado</th>

        <th>Recibido</th>

    </tr>


    {filas}


    </table>


    <table class="firmas">


    <tr>

        <th>

            ENTREGA

        </th>

        <th>

            RECEPCIÓN

        </th>

    </tr>


    <tr>


    <td class="firma-space">

        ____________________

        <br>

        {data["entregado_por"]}

    </td>


    <td class="firma-space">

        ____________________

        <br>

        {data["recibido_por"]}

        <br>

        {data["equipo_responsable"]}

    </td>


    </tr>


    </table>


    <div class="footer">

        Este documento certifica la entrega

        y recepción de los equipos

        y/o materiales descritos.

    </div>


    </div>


    </div>


    </body>


    </html>

    """


    return html


# ======================================================
# SIDEBAR
# ======================================================

with st.sidebar:


    if Path("logo.png").exists():

        st.image(

            "logo.png",

            use_container_width=True

        )


    st.markdown(

        '<div class="sidebar-btn sidebar-btn-active">'

        '📄 &nbsp; Nuevo Formulario'

        '</div>',

        unsafe_allow_html=True

    )


    st.markdown(

        '<div class="sidebar-btn">'

        '🕘 &nbsp; Historial de Entregas'

        '</div>',

        unsafe_allow_html=True

    )


    st.markdown(

        '<div class="sidebar-btn">'

        '📦 &nbsp; Catálogo de Materiales'

        '</div>',

        unsafe_allow_html=True

    )


    st.markdown(

        '<div class="sidebar-btn">'

        '📊 &nbsp; Reportes'

        '</div>',

        unsafe_allow_html=True

    )


    st.markdown(

        '<div class="sidebar-btn">'

        '⚙️ &nbsp; Configuración'

        '</div>',

        unsafe_allow_html=True

    )


    st.markdown(

        """

        <div class="help-box">

            <b>¿Necesitas ayuda?</b>

            <br>

            Contacta al administrador del sistema.

        </div>

        """,

        unsafe_allow_html=True

    )


# ======================================================
# ENCABEZADO
# ======================================================

numero_documento = obtener_siguiente_documento()

ahora = datetime.now()


st.markdown(

    '<h1 class="main-title">'

    'Entrega y Recepción de Equipos'

    '</h1>',

    unsafe_allow_html=True

)


st.markdown(

    '<p class="subtitle">'

    'Complete la información '

    'para generar el documento.'

    '</p>',

    unsafe_allow_html=True

)


# ======================================================
# ESTADO MATERIALES
# ======================================================

if "materiales" not in st.session_state:


    st.session_state.materiales = [


        {

            "Material / Equipo": "",

            "Cantidad": 1,

            "Unidad": "Unidad",

            "Entregado": True,

            "Recibido": True

        }


    ]


# ======================================================
# FORMULARIO
# ======================================================

left, right = st.columns(

    [1.05, 1.15],

    gap="large"

)


with left:


    st.markdown(

        '<div class="section-title">'

        '👷 1. DATOS GENERALES'

        '</div>',

        unsafe_allow_html=True

    )


    c1, c2 = st.columns(2)


    with c1:


        destino = st.text_input(

            "Destino / Área *"

        )


        proyecto = st.text_input(

            "Proyecto asignado *"

        )


        transportista = st.text_input(

            "Transportista *"

        )


        entregado_por = st.text_input(

            "Entregado por *",

            value="SOLARTEAM"

        )


    with c2:


        recibido_por = st.text_input(

            "Recibido por *"

        )


        equipo_responsable = st.text_input(

            "Equipo responsable"

        )


        fecha = st.date_input(

            "Fecha *",

            value=date.today()

        )


        observaciones = st.text_area(

            "Observaciones",

            height=125

        )


    st.divider()


    st.markdown(

        '<div class="section-title">'

        '📋 2. MATERIALES / EQUIPOS'

        '</div>',

        unsafe_allow_html=True

    )


    st.caption(

        "Puedes copiar desde Excel "

        "y pegar directamente en la tabla."

    )


    materiales = st.data_editor(

        st.session_state.materiales,

        num_rows="dynamic",

        use_container_width=True,

        hide_index=True,

        column_config={


            "Material / Equipo":

            st.column_config.TextColumn(

                "Material / Equipo"

            ),


            "Cantidad":

            st.column_config.NumberColumn(

                "Cantidad",

                min_value=0.0,

                step=1.0

            ),


            "Unidad":

            st.column_config.SelectboxColumn(

                "Unidad",

                options=UNIDADES

            ),


            "Entregado":

            st.column_config.CheckboxColumn(

                "Entregado"

            ),


            "Recibido":

            st.column_config.CheckboxColumn(

                "Recibido"

            )


        }

    )


    st.session_state.materiales = materiales


    data = {


        "numero_documento":

        numero_documento,


        "fecha":

        fecha.strftime("%d/%m/%Y"),


        "hora":

        ahora.strftime("%H:%M"),


        "destino":

        destino,


        "proyecto":

        proyecto,


        "transportista":

        transportista,


        "entregado_por":

        entregado_por,


        "recibido_por":

        recibido_por,


        "equipo_responsable":

        equipo_responsable,


        "observaciones":

        observaciones


    }


    materiales_validos = normalizar_materiales(

        materiales

    )


    generar = st.button(

        "⬇️ Generar PDF y Guardar",

        type="primary",

        use_container_width=True

    )


    if generar:


        campos_obligatorios = [

            destino,

            proyecto,

            transportista,

            entregado_por,

            recibido_por

        ]


        if not all(

            str(campo).strip()

            for campo

            in campos_obligatorios

        ):


            st.warning(

                "Completa destino, proyecto, "

                "transportista, entregado por "

                "y recibido por."

            )


        elif not materiales_validos:


            st.warning(

                "Agrega al menos un material "

                "o equipo."

            )


        else:


            pdf = generar_pdf(

                data,

                materiales_validos

            )


            pdf_name = (

                f"{numero_documento}.pdf"

            )


            pdf_path = (

                PDF_DIR / pdf_name

            )


            pdf_path.write_bytes(

                pdf.getvalue()

            )


            guardar_excel(

                data,

                materiales_validos,

                str(pdf_path)

            )


            st.success(

                f"Formulario "

                f"{numero_documento} "

                f"generado correctamente."

            )


            st.download_button(

                "Descargar PDF",

                pdf.getvalue(),

                pdf_name,

                "application/pdf",

                use_container_width=True

            )


# ======================================================
# VISTA PREVIA
# ======================================================

with right:


    st.markdown(

        '<div class="section-title">'

        'VISTA PREVIA DEL DOCUMENTO'

        '</div>',

        unsafe_allow_html=True

    )


    components.html(

        pdf_preview_html(

            data,

            materiales_validos

        ),

        height=900,

        scrolling=True

    )


# ======================================================
# HISTORIAL
# ======================================================

st.markdown(

    '<div class="history-card">',

    unsafe_allow_html=True

)


hcol1, hcol2 = st.columns(

    [4, 1]

)


with hcol1:


    st.markdown(

        '<div class="section-title">'

        '🧾 HISTORIAL DE ENTREGAS'

        '</div>',

        unsafe_allow_html=True

    )


with hcol2:


    if EXCEL_FILE.exists():


        with open(

            EXCEL_FILE,

            "rb"

        ) as archivo_excel:


            st.download_button(

                "📗 Exportar Excel",

                archivo_excel,

                "registro_entregas.xlsx",

                (
                    "application/"
                    "vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                )

            )


historial = cargar_historial()


if historial.empty:


    st.info(

        "Todavía no existen "

        "entregas registradas."

    )


else:


    busqueda = st.text_input(

        "Buscar por proyecto, destino "

        "o N° de documento"

    )


    if busqueda.strip():


        filtro = (

            historial

            .astype(str)

            .apply(

                lambda fila:

                fila.str.contains(

                    busqueda,

                    case=False,

                    na=False

                ).any(),

                axis=1

            )

        )


        historial = historial[

            filtro

        ]


    st.dataframe(

        historial,

        use_container_width=True,

        hide_index=True

    )


st.markdown(

    '</div>',

    unsafe_allow_html=True

)
